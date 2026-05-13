# Google SERP Automation 
import random
import time
import pandas as pd
import urllib.parse
import asyncio
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from crawl4ai import AsyncWebCrawler

# Define input and output file paths
input_path = "/path/to/your/input_file.xlsx"
output_path = "/path/to/your/output_file.xlsx"

# Load all sheets
all_sheets = pd.read_excel(input_path, sheet_name=None) # Dict of DataFrames

# Async function to check for sponsored content
async def check_sponsored(query):
    async with AsyncWebCrawler() as crawler:
        url = f"https://www.google.com/search?q={urllib.parse.quote(str(query))}"
        result = await crawler.arun(url=url)
        if hasattr(result, 'markdown'):
            return result.markdown.lower().count('sponsored')
    return 0

# Async function to process queries in a DataFrame
async def process_sheet(df):
    sponsored_counts = []
    sponsored_presence = []
    for query in df["Query"]:
        sponsored_count = await check_sponsored(query)
        sponsored_counts.append(sponsored_count)
        sponsored_presence.append("TRUE" if sponsored_count > 0 else "FALSE")
        await asyncio.sleep(random.uniform(1, 3)) # Random delay
    return sponsored_counts, sponsored_presence

# Prepare output dictionary for all sheets
output_sheets = {}

# Process each sheet
for sheet_name, df in all_sheets.items():
    print(f"Processing sheet: {sheet_name}")
    if "Query" not in df.columns:
        print(f"Skipping '{sheet_name}' (no 'Query' column)")
        continue

    df["Search URL"] = df["Query"].apply(lambda q: f"https://www.google.com/search?q={urllib.parse.quote(str(q))}")
    sponsored_counts, sponsored_presence = asyncio.run(process_sheet(df))
    df["Sponsored Count"] = sponsored_counts
    df["Sponsored Presence"] = sponsored_presence
    output_sheets[sheet_name] = df

# Write all sheets to new Excel file
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    for sheet_name, df in output_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False

# Add hyperlinks using openpyxl

wb = load_workbook(output_path)



# Define hyperlink style

hyperlink_style = NamedStyle(name="Hyperlink")

hyperlink_style.font = hyperlink_style.font.copy(underline="single", color="0563C1")

if "Hyperlink" not in wb.named_styles:

wb.add_named_style(hyperlink_style)



# Apply hyperlinks to each sheet

for sheet_name in output_sheets.keys():
    ws = wb[sheet_name]

for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=3) # Column 3 = 'Search URL'
    url = cell.value

if url:
    cell.hyperlink = url
    cell.value = "GG"
    cell.style = "Hyperlink"

wb.save(output_path)

print(f"Hyperlinked Excel saved to: {output_path}")